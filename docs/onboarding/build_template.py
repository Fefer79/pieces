#!/usr/bin/env python3
"""Génère le modèle Excel d'onboarding d'une flotte Pièces.
Champs alignés sur le schéma Prisma (Enterprise / Vehicle / Driver)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

NAVY = "00113A"
ORANGE = "FF6B00"
LIGHT = "F2F4F8"
WHITE = "FFFFFF"

navy_fill = PatternFill("solid", fgColor=NAVY)
orange_fill = PatternFill("solid", fgColor=ORANGE)
light_fill = PatternFill("solid", fgColor=LIGHT)
header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
title_font = Font(name="Calibri", bold=True, color=NAVY, size=18)
sub_font = Font(name="Calibri", color=NAVY, size=11)
label_font = Font(name="Calibri", bold=True, color=NAVY, size=11)
hint_font = Font(name="Calibri", italic=True, color="666666", size=9)
thin = Side(style="thin", color="D0D5DD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center")

wb = Workbook()

# ---------------------------------------------------------------- Instructions
ws = wb.active
ws.title = "Mode d'emploi"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 100
ws["B2"] = "Pièces — Fiche d'onboarding flotte"
ws["B2"].font = title_font
ws["B3"] = "Modèle à remplir pour enregistrer une nouvelle entreprise et ses véhicules sur la plateforme."
ws["B3"].font = sub_font
steps = [
    ("Comment l'utiliser", ""),
    ("1.", "Remplir l'onglet « Entreprise » (une seule entreprise par fichier)."),
    ("2.", "Lister les véhicules dans l'onglet « Véhicules », une ligne par véhicule."),
    ("3.", "Lister les chauffeurs dans l'onglet « Chauffeurs » (facultatif)."),
    ("4.", "Renvoyer le fichier à l'équipe Pièces ou l'importer depuis le tableau de bord entreprise."),
    ("", ""),
    ("Règles de saisie", ""),
    ("•", "Téléphone au format ivoirien : +225 suivi de 10 chiffres (ex : +2250700000000)."),
    ("•", "Plaque d'immatriculation : format CI (ex : 1234 AB 01)."),
    ("•", "Année : 4 chiffres (ex : 2018). Kilométrage : en km, chiffres uniquement."),
    ("•", "Les colonnes avec une liste déroulante : choisir une valeur proposée."),
    ("•", "Colonnes marquées * = obligatoires. Ne pas renommer les en-têtes."),
]
r = 5
for tag, txt in steps:
    if txt == "" and tag and not tag[0].isdigit() and tag not in ("•",):
        ws[f"B{r}"] = tag
        ws[f"B{r}"].font = Font(bold=True, color=ORANGE, size=12)
    else:
        ws[f"B{r}"] = (f"{tag}  {txt}" if tag else txt).strip()
        ws[f"B{r}"].font = sub_font
        ws[f"B{r}"].alignment = wrap
    r += 1

def style_header(ws, headers, hints, row=1):
    """En-têtes en ligne 1 ; les indices de saisie sont des commentaires de
    cellule (survol) pour que la 1re ligne de données reste la ligne 2 —
    importable sans ligne parasite."""
    for c, (h, hint) in enumerate(zip(headers, hints), start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = navy_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = border
        if hint:
            cm = Comment(hint, "Pièces")
            cm.width, cm.height = 200, 60
            cell.comment = cm
    ws.row_dimensions[row].height = 30

# ---------------------------------------------------------------- Entreprise
we = wb.create_sheet("Entreprise")
we.sheet_view.showGridLines = False
we.column_dimensions["A"].width = 32
we.column_dimensions["B"].width = 45
we.column_dimensions["C"].width = 50
we["A1"] = "Champ"
we["B1"] = "Valeur à remplir"
we["C1"] = "Aide"
for col in ("A1", "B1", "C1"):
    we[col].fill = navy_fill
    we[col].font = header_font
    we[col].border = border
ent_rows = [
    ("Nom de l'entreprise *", "", "Raison sociale telle qu'elle apparaîtra sur les factures."),
    ("Commune", "", "Ex : Yopougon, Cocody, Marcory…"),
    ("Adresse", "", "Adresse physique / point de repère."),
    ("RCCM", "", "Numéro d'immatriculation au registre du commerce (si dispo)."),
    ("Nom du contact principal *", "", "Personne responsable du compte."),
    ("Téléphone du contact *", "", "+225XXXXXXXXXX — sert d'accès OTP au tableau de bord."),
    ("Email du contact", "", "Pour l'envoi des factures et reçus."),
    ("Rôle du contact", "", "OWNER / MANAGER / ACCOUNTANT."),
    ("Formule souhaitée", "", "Gratuit / Flotte Pro / Flotte Pro +."),
    ("Nombre de véhicules", "", "Total approximatif de la flotte."),
]
for i, (label, val, hint) in enumerate(ent_rows, start=2):
    we.cell(row=i, column=1, value=label).font = label_font
    we.cell(row=i, column=1).border = border
    we.cell(row=i, column=1).fill = light_fill
    vc = we.cell(row=i, column=2, value=val)
    vc.border = border
    hc = we.cell(row=i, column=3, value=hint)
    hc.font = hint_font
    hc.alignment = wrap
    hc.border = border
    we.row_dimensions[i].height = 22

# dropdowns entreprise
dv_role = DataValidation(type="list", formula1='"OWNER,MANAGER,ACCOUNTANT"', allow_blank=True)
dv_plan = DataValidation(type="list", formula1='"Gratuit,Flotte Pro,Flotte Pro +"', allow_blank=True)
we.add_data_validation(dv_role); dv_role.add(we["B9"])
we.add_data_validation(dv_plan); dv_plan.add(we["B10"])

# ---------------------------------------------------------------- Véhicules
wv = wb.create_sheet("Véhicules")
wv.sheet_view.showGridLines = False
veh_headers = ["Marque *", "Modèle *", "Année *", "Immatriculation", "VIN",
               "Motorisation", "Kilométrage", "Type d'usage", "Groupe / Site", "Chauffeur attitré"]
veh_hints = ["Toyota, Hyundai…", "Hilux, H1…", "2018", "1234 AB 01", "N° de châssis (17 car.)",
             "2.5 D-4D, essence…", "en km", "voir liste", "ex : Dépôt Nord", "nom du chauffeur"]
style_header(wv, veh_headers, veh_hints)
widths = [16, 16, 8, 16, 20, 16, 12, 18, 18, 20]
for i, w in enumerate(widths, start=1):
    wv.column_dimensions[get_column_letter(i)].width = w
# borders for entry rows (data starts row 2)
for row in range(2, 60):
    for c in range(1, len(veh_headers) + 1):
        wv.cell(row=row, column=c).border = border
dv_usage = DataValidation(type="list",
    formula1='"TRANSPORT,CHANTIER,LIVRAISON,DIRECTION,AUTRE"', allow_blank=True)
wv.add_data_validation(dv_usage)
dv_usage.add("H2:H59")
dv_year = DataValidation(type="whole", operator="between", formula1="1980", formula2="2030", allow_blank=True)
wv.add_data_validation(dv_year); dv_year.add("C2:C59")
wv.freeze_panes = "A2"

# ---------------------------------------------------------------- Chauffeurs
wd = wb.create_sheet("Chauffeurs")
wd.sheet_view.showGridLines = False
dr_headers = ["Nom complet *", "Téléphone *", "N° de permis", "Catégorie permis", "Date d'embauche", "Notes"]
dr_hints = ["Prénom Nom", "+2250700000000", "", "B, C, D…", "JJ/MM/AAAA", "remarques"]
style_header(wd, dr_headers, dr_hints)
for i, w in enumerate([24, 20, 18, 16, 16, 34], start=1):
    wd.column_dimensions[get_column_letter(i)].width = w
for row in range(2, 40):
    for c in range(1, len(dr_headers) + 1):
        wd.cell(row=row, column=c).border = border
wd.freeze_panes = "A2"

outs = [
    "/Users/mac/dev/pieces/docs/onboarding/modele-onboarding-flotte.xlsx",
    # Servi en téléchargement depuis la page d'import du tableau de bord.
    "/Users/mac/dev/pieces/apps/web/public/modele-onboarding-flotte.xlsx",
]
for out in outs:
    wb.save(out)
    print("saved", out)
